import csv
import io
from dataclasses import dataclass, field
from typing import List, Optional

from openpyxl import Workbook, load_workbook

from app.schemas.proveedores import (
    ContactoProveedorEmbedded,
    EmailContactoEmbedded,
    ProveedorCreate,
    TelefonoContactoEmbedded,
)

COLUMNAS = [
    "nit",
    "tipo_persona",
    "razon_social",
    "nombre_comercial",
    "pais",
    "idioma",
    "sitio_web",
    "moneda_defecto",
    "notas",
    "contacto_nombre",
    "contacto_cargo",
    "contacto_email",
    "contacto_telefono",
]

FILA_EJEMPLO = [
    "900123456-7",
    "juridica",
    "Proveedor Ejemplo S.A.S.",
    "Proveedor Ejemplo",
    "Colombia",
    "ES",
    "https://proveedorejemplo.com",
    "COP",
    "Proveedor de referencia cargado como ejemplo.",
    "Juan Pérez",
    "Gerente Comercial",
    "juan.perez@proveedorejemplo.com",
    "3001234567",
]


def generar_plantilla() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    ws.append(COLUMNAS)
    ws.append(FILA_EJEMPLO)
    for col_idx, header in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(18, len(header) + 4)

    instrucciones = wb.create_sheet("Instrucciones")
    filas_instrucciones = [
        ["Campo", "Obligatorio", "Descripción"],
        ["nit", "No", "Número de identificación tributaria. Si ya existe un proveedor con este NIT, la fila se omite."],
        ["tipo_persona", "No", "juridica o natural."],
        ["razon_social", "Sí", "Nombre legal del proveedor."],
        ["nombre_comercial", "No", "Nombre comercial o de fantasía."],
        ["pais", "Sí", "País del proveedor."],
        ["idioma", "No", "ES o EN. Si se deja vacío y el país es Colombia, se asigna ES automáticamente."],
        ["sitio_web", "No", "URL del sitio web del proveedor."],
        ["moneda_defecto", "No", "Moneda principal: COP, USD, EUR, etc."],
        ["notas", "No", "Notas internas sobre el proveedor."],
        ["contacto_nombre", "No", "Nombre del contacto principal. Si se completa, se crea un contacto asociado."],
        ["contacto_cargo", "No", "Cargo del contacto."],
        ["contacto_email", "No", "Correo del contacto."],
        ["contacto_telefono", "No", "Teléfono del contacto."],
        [],
        ["Nota: cada fila representa un proveedor con, como máximo, un contacto. Para agregar más "
         "contactos a un proveedor, hacelo luego desde el módulo de Proveedores."],
    ]
    for fila in filas_instrucciones:
        instrucciones.append(fila)
    instrucciones.column_dimensions["A"].width = 20
    instrucciones.column_dimensions["B"].width = 14
    instrucciones.column_dimensions["C"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@dataclass
class FilaError:
    fila: int
    motivo: str


@dataclass
class ResultadoImportacion:
    creados: int = 0
    omitidos_duplicados: List[str] = field(default_factory=list)
    errores: List[FilaError] = field(default_factory=list)


def _leer_filas(contenido: bytes, nombre_archivo: str) -> List[dict]:
    nombre = nombre_archivo.lower()
    if nombre.endswith(".csv"):
        texto = contenido.decode("utf-8-sig")
        lector = csv.DictReader(io.StringIO(texto))
        return [dict(fila) for fila in lector]

    if nombre.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb["Proveedores"] if "Proveedores" in wb.sheetnames else wb.worksheets[0]
        filas_iter = ws.iter_rows(values_only=True)
        encabezados = [str(h).strip() if h is not None else "" for h in next(filas_iter)]
        filas = []
        for valores in filas_iter:
            if valores is None or all(v is None for v in valores):
                continue
            filas.append({
                encabezados[i]: valores[i] if i < len(valores) else None
                for i in range(len(encabezados))
            })
        return filas

    raise ValueError("Formato de archivo no soportado. Usá .csv o .xlsx.")


def _valor(fila: dict, clave: str) -> Optional[str]:
    v = fila.get(clave)
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def _construir_proveedor(fila: dict) -> ProveedorCreate:
    razon_social = _valor(fila, "razon_social")
    pais = _valor(fila, "pais")
    if not razon_social:
        raise ValueError("razon_social es obligatorio.")
    if not pais:
        raise ValueError("pais es obligatorio.")

    contactos: List[ContactoProveedorEmbedded] = []
    contacto_nombre = _valor(fila, "contacto_nombre")
    if contacto_nombre:
        emails = []
        contacto_email = _valor(fila, "contacto_email")
        if contacto_email:
            emails.append(EmailContactoEmbedded(email=contacto_email))
        telefonos = []
        contacto_telefono = _valor(fila, "contacto_telefono")
        if contacto_telefono:
            telefonos.append(TelefonoContactoEmbedded(numero=contacto_telefono))
        contactos.append(ContactoProveedorEmbedded(
            nombre=contacto_nombre,
            cargo=_valor(fila, "contacto_cargo"),
            es_principal=True,
            emails=emails,
            telefonos=telefonos,
        ))

    return ProveedorCreate(
        nit=_valor(fila, "nit"),
        tipo_persona=_valor(fila, "tipo_persona"),
        razon_social=razon_social,
        nombre_comercial=_valor(fila, "nombre_comercial"),
        pais=pais,
        idioma=_valor(fila, "idioma") or "ES",
        sitio_web=_valor(fila, "sitio_web"),
        notas=_valor(fila, "notas"),
        moneda_defecto=_valor(fila, "moneda_defecto"),
        contactos=contactos,
    )


def parsear_archivo(contenido: bytes, nombre_archivo: str) -> List[dict]:
    return _leer_filas(contenido, nombre_archivo)


def construir_proveedor_desde_fila(fila: dict) -> ProveedorCreate:
    return _construir_proveedor(fila)
